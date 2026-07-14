"""④ 판매(결제)대행 매출 엑셀 후가공 — 상호별 정리본 생성.

홈택스 원본(.xls, 승인년월순 나열)을 읽어:
    Sheet1(정리본): 판매(결제)대행업체 상호 가나다순으로 그룹핑,
                    그룹마다 승인년월순 나열 + 금액 3열 소계, 맨 아래 총합계.
                    상단에 '< 1기확정 - 업체명 >' 제목.
    Sheet2(원본):   내려받은 표 그대로.
형식은 사용자가 수작업으로 만들던 정리본(2026-01 예시 파일)을 그대로 재현.

원본 .xls 구조 (2026-07 확인):
    row0 빈 줄 / row1-2 이중 헤더(순번|승인년월|건수|매출금액×3|상호,
    2행째 D~F가 계|신용카드|기타결제수단) / row3 합계 행 / row4~ 데이터.
"""
from __future__ import annotations

from pathlib import Path


def make_summary(src_path, title: str, log=print) -> tuple[bool, str, str]:
    """원본 .xls → 정리본 .xlsx (Sheet1 정리본 / Sheet2 원본).

    return: (성공, 사유, 생성 파일 경로). 실패 시 원본을 그대로 두고 (False, 사유, "").
    """
    src = Path(src_path)
    try:
        import xlrd
        wb = xlrd.open_workbook(str(src))
        sh = wb.sheet_by_index(0)
        grid = [[sh.cell_value(r, c) for c in range(sh.ncols)]
                for r in range(sh.nrows)]
    except Exception as e:
        return False, f"원본 읽기 실패: {str(e)[:80]}", ""

    # ── 데이터 행 파싱: '합계' 행 다음부터, 승인년월(YYYY-MM)이 있는 행 ──
    rows = []
    started = False
    for row in grid:
        c0 = str(row[0]).strip() if row else ""
        if not started:
            if c0 == "합계":
                started = True
            continue
        ym = str(row[1]).strip() if len(row) > 1 else ""
        if len(ym) != 7 or ym[4] != "-":
            continue
        def num(i):
            try:
                return int(float(row[i] or 0))
            except Exception:
                return 0
        name = str(row[6]).strip() if len(row) > 6 else ""
        rows.append({"ym": ym, "cnt": num(2), "tot": num(3),
                     "card": num(4), "etc": num(5), "agency": name})
    if not rows:
        return False, "데이터 행을 찾지 못함 (원본 형식 변경?)", ""

    # ── 상호 가나다순 그룹핑, 그룹 내 승인년월순 ──
    agencies = sorted({r["agency"] for r in rows})
    groups = [(a, sorted((r for r in rows if r["agency"] == a),
                         key=lambda r: r["ym"])) for a in agencies]

    # ── xlsx 작성 ──
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        out = openpyxl.Workbook()
        ws = out.active
        ws.title = "정리본"

        thin = Side(style="thin", color="999999")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        head_fill = PatternFill("solid", fgColor="DDEBF7")
        sub_fill = PatternFill("solid", fgColor="F2F2F2")
        bold = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        NUM = "#,##0"

        def put(r, c, v, *, b=False, fill=None, num=False, cen=False):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = border
            if b:
                cell.font = bold
            if fill:
                cell.fill = fill
            if num:
                cell.number_format = NUM
            if cen:
                cell.alignment = center
            return cell

        # 제목 (예시 형식: '< 2기확정 - 업체명 >' — E열)
        tcell = ws.cell(row=2, column=5, value=f"< {title} >")
        tcell.font = Font(bold=True, size=12)

        # 이중 헤더 (4~5행) + 병합
        heads1 = ["순번", "승인년월", "건수", "매출금액", "매출금액", "매출금액",
                  "판매(결제)대행업체 상호"]
        heads2 = ["순번", "승인년월", "건수", "계", "신용카드", "기타결제수단",
                  "판매(결제)대행업체 상호"]
        for c, v in enumerate(heads1, 1):
            put(4, c, v, b=True, fill=head_fill, cen=True)
        for c, v in enumerate(heads2, 1):
            put(5, c, v, b=True, fill=head_fill, cen=True)
        ws.merge_cells("A4:A5")
        ws.merge_cells("B4:B5")
        ws.merge_cells("C4:C5")
        ws.merge_cells("D4:F4")
        ws.merge_cells("G4:G5")

        # 상호별 그룹 + 소계 — 합계는 값이 아니라 살아있는 수식으로 (사용자 요청)
        r_i, seq = 7, 1
        sub_rows = []          # 상호별 소계 행 번호들
        for agency, g in groups:
            g_start = r_i
            for row in g:
                put(r_i, 1, seq, cen=True)
                put(r_i, 2, row["ym"], cen=True)
                put(r_i, 3, row["cnt"], num=True)
                put(r_i, 4, row["tot"], num=True)
                put(r_i, 5, row["card"], num=True)
                put(r_i, 6, row["etc"], num=True)
                put(r_i, 7, agency)
                r_i += 1
                seq += 1
            g_end = r_i - 1
            put(r_i, 1, "")
            put(r_i, 2, "")
            put(r_i, 3, "")
            for col, letter in ((4, "D"), (5, "E"), (6, "F")):
                put(r_i, col, f"=SUM({letter}{g_start}:{letter}{g_end})",
                    b=True, fill=sub_fill, num=True)
            put(r_i, 7, agency, b=True, fill=sub_fill)
            sub_rows.append(r_i)
            r_i += 1

        # 전체 합계 행 (6행) — 건수는 데이터 구간 SUM(소계 행 C는 빈칸이라 안전),
        # 금액은 상호별 소계 셀들의 합
        last_sub = sub_rows[-1]
        put(6, 1, "합계", b=True, fill=head_fill, cen=True)
        put(6, 2, "합계", b=True, fill=head_fill, cen=True)
        put(6, 3, f"=SUM(C7:C{last_sub})", b=True, fill=head_fill, num=True)
        for col, letter in ((4, "D"), (5, "E"), (6, "F")):
            expr = "+".join(f"{letter}{r}" for r in sub_rows)
            put(6, col, f"={expr}", b=True, fill=head_fill, num=True)
        put(6, 7, "", fill=head_fill)

        # 총합계 행 — 상호별 소계 셀들의 합 (상단 합계와 동일 수식)
        for col, letter in ((4, "D"), (5, "E"), (6, "F")):
            expr = "+".join(f"{letter}{r}" for r in sub_rows)
            put(r_i, col, f"={expr}", b=True, num=True)

        for col, w in zip("ABCDEFG", (7, 11, 9, 13, 13, 13, 28)):
            ws.column_dimensions[col].width = w

        # Sheet2: 원본 그대로
        ws2 = out.create_sheet("원본")
        for r, row in enumerate(grid, 1):
            for c, v in enumerate(row, 1):
                if isinstance(v, float) and v == int(v):
                    v = int(v)
                ws2.cell(row=r, column=c, value=v)

        out_path = src.with_suffix(".xlsx")
        if out_path.exists():
            out_path.unlink()
        out.save(out_path)
    except Exception as e:
        return False, f"정리본 작성 실패: {str(e)[:80]}", ""

    try:
        src.unlink()   # 원본 .xls는 Sheet2에 들어갔으므로 파일은 정리
    except Exception:
        pass
    log(f"    엑셀 정리본 생성: 상호 {len(groups)}곳, {len(rows)}행 → {out_path.name}")
    return True, "", str(out_path)
