"""실행 결과 엑셀 — 업체 × 작업 결과표.

사업자번호 오류로 중단/건너뛴 업체를 사용자가 한눈에 보고 명부를 고칠 수 있게 남긴다.
저장 위치: output_dir (미지정 시 앱 데이터 폴더), 파일명에 실행 시각 포함.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from .browser import app_data_dir
from .util import fmt_bizno


def _status(r) -> str:
    if r.ok:
        return "성공"
    if r.skipped:
        return "건너뜀"
    if r.fatal:
        return "사업자번호 오류"
    return "실패"


def write_results(results: list, clients: list[dict], inp) -> str:
    """PhaseResult 리스트를 엑셀로 저장하고 경로 반환."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    name_to_bizno = {c.get("name", ""): c.get("bizno", "") for c in clients}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "조회결과"
    ws.append([f"부가세 신고자료 출력 결과 — {inp.year}년 {inp.term}기"
               f" ({datetime.now():%Y-%m-%d %H:%M})"])
    ws.append([])
    header = ["업체명", "사업자등록번호", "작업", "결과", "사유", "산출물"]
    ws.append(header)
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="E2E8F0")
    for cell in ws[3]:
        cell.font = bold
        cell.fill = fill

    red = Font(color="B91C1C")
    for r in results:
        row = [r.client_name, fmt_bizno(name_to_bizno.get(r.client_name, "")),
               r.label, _status(r), r.reason, "; ".join(r.outputs)]
        ws.append(row)
        if not r.ok:
            ws.cell(row=ws.max_row, column=4).font = red

    for col, w in zip("ABCDEF", (24, 16, 24, 14, 50, 60)):
        ws.column_dimensions[col].width = w

    out_dir = Path(inp.output_dir) if inp.output_dir else app_data_dir()
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"조회결과_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    wb.save(path)
    return str(path)
